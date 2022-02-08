from django.db import models
from django.shortcuts import render
from sales.models import OrdersDetail, OrdersHeader
# import xlwings as xw
from openpyxl import Workbook
import datetime
from django.shortcuts import HttpResponse
from openpyxl.writer.excel import save_virtual_workbook


# Create your views here.


def download_orders_view(request):
    # response = HttpResponse(content_type='application/ms-excel')
    wk = Workbook()
    st = wk.active
    st.title = "客户订单"
    columns = [
        '订单日期',
        '订单号',
        '客户姓名',
        '款号',
        '颜色',
        '数量',
        '单价',
        '总价',
        '订单总价',
        '电话',
        '已发货',
    ]
    # st['A1:K1'] = columns
    st.append(columns)
    orders_list = OrdersHeader.objects.all()
    for order_header in orders_list:
        orders_detail_list = OrdersDetail.objects.filter(order_header=order_header)
        # start_row = st.range("D999999").end('up')
        start_row = st.max_row
        start_lst = [order_header.order_date, order_header.order_num,order_header.customer.name]
        end_lst = [order_header.order_price,order_header.customer.phone, order_header.issued_all]
        for order_detail in orders_detail_list:
            # row_num = start_row.offset(1).row
            row_num = start_row + 1
            middle_lst = [order_detail.clothe_num,order_detail.color,order_detail.amount,order_detail.price]
            start_lst.extend(middle_lst)
            start_lst.extend(end_lst)
            # st['A%s:K%s'%(row_num,row_num)].value = start_lst
            st.append(start_lst)
            # start_num = start_row.offset(1).row + 1
        start_num = start_row + 1
        end_num = start_num + len(orders_detail_list) - 1
        # st.range('A%s:A%s'%(start_num,end_num)).merge()
        # st.range('B%s:B%s'%(start_num,end_num)).merge()
        # st.range('C%s:C%s'%(start_num,end_num)).merge()
        # st.range('I%s:I%s'%(start_num,end_num)).merge()
        # st.range('J%s:J%s'%(start_num,end_num)).merge()
        # st.range('K%s:K%s'%(start_num,end_num)).merge()
        st.merge_cells('A%s:A%s'%(start_num,end_num))
        st.merge_cells('B%s:B%s'%(start_num,end_num))
        st.merge_cells('C%s:C%s'%(start_num,end_num))
        st.merge_cells('I%s:I%s'%(start_num,end_num))
        st.merge_cells('J%s:J%s'%(start_num,end_num))
        st.merge_cells('K%s:K%s'%(start_num,end_num))
    wk.close()
    response = HttpResponse(save_virtual_workbook(wk), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['content-Disposition'] = 'attachment; filename=订单' + str(datetime.datetime.now()) + '.xlsx'

    wk.save(response)
    return response

def generate_excel():
    pass